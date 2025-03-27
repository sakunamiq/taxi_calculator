import tkinter as tk
from tkinter import ttk, messagebox
import math
import os
import datetime
try:
    import openpyxl
    from tkcalendar import Calendar, DateEntry
except ImportError:
    missing_packages = []
    try:
        import openpyxl
    except ImportError:
        missing_packages.append("openpyxl")
    try:
        from tkcalendar import Calendar, DateEntry
    except ImportError:
        missing_packages.append("tkcalendar")
    
    if missing_packages:
        error_msg = f"Пожалуйста, установите необходимые пакеты: {', '.join(missing_packages)}\nКоманда для установки: pip install {' '.join(missing_packages)}"
        messagebox.showerror("Ошибка импорта", error_msg)
        exit(1)

EXCEL_FILENAME = "taxi_payment_log.xlsx"

MAIN_BG = "#f0f0f0"
HEADER_BG = "#2c3e50"
HEADER_FG = "white"
BUTTON_BG = "#3498db"
BUTTON_FG = "white"
ACCENT_COLOR = "#e74c3c"
ENTRY_BG = "white"
RESULT_BG = "#ecf0f1"

def calculate_gasoline_expense(total_revenue):
    if 3000 <= total_revenue < 4000:
        return 500
    elif 4000 <= total_revenue < 4500:
        return 800
    elif 4500 <= total_revenue < 5000:
        return 850
    elif 5000 <= total_revenue < 5500:
        return 900
    elif 5500 <= total_revenue < 6000:
        return 1000
    elif 6000 <= total_revenue < 6500:
        return 1050
    elif 6500 <= total_revenue < 7000:
        return 1100
    elif 7000 <= total_revenue < 7500:
        return 1150
    elif 7500 <= total_revenue < 8000:
        return 1200
    elif 8000 <= total_revenue < 8500:
        return 1250
    elif 8500 <= total_revenue < 9000:
        return 1300
    elif 9000 <= total_revenue < 9500:
        return 1350
    elif 9500 <= total_revenue < 10000:
        return 1400
    elif 10000 <= total_revenue < 10500:
        return 1450
    elif 10500 <= total_revenue < 11000:
        return 1500
    elif 11000 <= total_revenue < 11500:
        return 1550
    elif 11500 <= total_revenue < 12000:
        return 1600
    elif 12000 <= total_revenue < 12500:
        return 1650
    elif 12500 <= total_revenue < 13000:
        return 1700
    elif 13000 <= total_revenue < 13500:
        return 1750
    elif 13500 <= total_revenue < 14000:
        return 1800
    elif 14000 <= total_revenue < 14500:
        return 1850
    elif 14500 <= total_revenue < 15000:
        return 1900
    elif total_revenue >= 15000:
        return 1900
    else:
        return 0

def validate_number_input(text):
    if text == "":
        return True
    try:
        float(text)
        return True
    except ValueError:
        return False

def validate_date_input(text):
    if text == "":
        return True
    if not all(c.isdigit() or c == '.' for c in text):
        return False
    if len(text) > 10:
        return False
    return True

def open_calendar():
    def get_date():
        selected_date = cal.get_date()
        formatted_date = selected_date.strftime("%d.%m.%Y")
        date_var.set(formatted_date)
        top.destroy()
    
    today = datetime.date.today()
    top = tk.Toplevel(window)
    top.title("Выберите дату")
    top.grab_set()  # Make the popup modal
    
    cal_frame = tk.Frame(top, padx=10, pady=10)
    cal_frame.pack(fill="both", expand=True)
    
    cal = Calendar(cal_frame, selectmode="day", year=today.year, month=today.month, day=today.day,
                  locale='ru_RU', background=MAIN_BG, foreground="black", 
                  selectbackground=ACCENT_COLOR, selectforeground="white",
                  normalbackground=ENTRY_BG, weekendbackground="#d9d9d9")
    cal.pack(fill="both", expand=True, padx=10, pady=10)
    
    btn_frame = tk.Frame(top, padx=10, pady=5)
    btn_frame.pack(fill="x")
    
    select_btn = tk.Button(btn_frame, text="Выбрать", command=get_date, 
                          bg=BUTTON_BG, fg=BUTTON_FG, padx=10)
    select_btn.pack(side="right", padx=5)
    
    cancel_btn = tk.Button(btn_frame, text="Отмена", command=top.destroy, 
                          bg="#7f8c8d", fg=BUTTON_FG, padx=10)
    cancel_btn.pack(side="right", padx=5)

def calculate_payment_gui():
    try:
        if not fio_entry.get().strip():
            messagebox.showwarning("Предупреждение", "Пожалуйста, введите ФИО водителя")
            return
        
        if not date_var.get().strip():
            messagebox.showwarning("Предупреждение", "Пожалуйста, введите дату")
            return
            
        if not revenue_entry.get().strip():
            messagebox.showwarning("Предупреждение", "Пожалуйста, введите общую выручку")
            return
            
        if not commission_entry.get().strip():
            messagebox.showwarning("Предупреждение", "Пожалуйста, введите сумму комиссии")
            return
        
        total_revenue = float(revenue_entry.get())
        commission = float(commission_entry.get())
        driver_fio = fio_entry.get()
        date_value = date_var.get()
        car_model = car_combobox.get()

        gasoline = calculate_gasoline_expense(total_revenue)

        labor_percentage = 0
        if 3000 <= total_revenue < 4500:
            labor_percentage = 42
        elif 4500 <= total_revenue < 5000:
            labor_percentage = 42
        elif 5000 <= total_revenue < 6000:
            labor_percentage = 43
        elif 6000 <= total_revenue < 7000:
            labor_percentage = 45
        elif 7000 <= total_revenue < 8000:
            labor_percentage = 46
        elif 8000 <= total_revenue < 9000:
            labor_percentage = 47
        elif 9000 <= total_revenue < 10000:
            labor_percentage = 48
        elif total_revenue >= 10000:
            labor_percentage = 50

        labor_payment_decimal = total_revenue * (labor_percentage / 100)

        weekend_worked = weekend_var.get()
        weekend_bonus_decimal = 0
        if weekend_worked:
            weekend_bonus_decimal = total_revenue * 0.05

        top_place_bonus_percentage = 0
        top_place = top_place_var.get()
        if top_place == 1:
            top_place_bonus_percentage = 5
        elif top_place == 2:
            top_place_bonus_percentage = 3
        elif top_place == 3:
            top_place_bonus_percentage = 1

        top_place_bonus_decimal = total_revenue * (top_place_bonus_percentage / 100)

        driver_payment_decimal = labor_payment_decimal + weekend_bonus_decimal + top_place_bonus_decimal
        driver_payment = math.floor(driver_payment_decimal)
        labor_payment = math.floor(labor_payment_decimal)  
        weekend_bonus = math.floor(weekend_bonus_decimal)  
        top_place_bonus = math.floor(top_place_bonus_decimal)

        fleet_payment = total_revenue - driver_payment - gasoline - commission

        result_frame.config(bg=RESULT_BG)
        result_text = f"Результаты расчета:"
        result_label.config(text=result_text, bg=RESULT_BG, fg=HEADER_BG, font=("Arial", 12, "bold"))
        
        results = {
            "Общая выручка": f"{total_revenue:.2f} руб.",
            "Ручная комиссия": f"{commission:.2f} руб.",
            "Трудовая выплата (авто-расчет)": f"{labor_payment_decimal:.2f} руб. ({labor_percentage}%)",
            "Расходы на бензин (авто-расчет)": f"{gasoline:.2f} руб.",
            "Бонус за выходные": f"{weekend_bonus_decimal:.2f} руб.",
            "Бонус за топ место": f"{top_place_bonus_decimal:.2f} руб. ({top_place_bonus_percentage}%)",
            "Общая выплата водителю": f"{driver_payment_decimal:.2f} руб.",
            "Выплата водителю": f"{driver_payment:.2f} руб.",
            "Выплата парку": f"{fleet_payment:.2f} руб."
        }
        
        for widget in result_details_frame.winfo_children():
            widget.destroy()
            
        row = 0
        for key, value in results.items():
            label = tk.Label(result_details_frame, text=key, anchor="w", bg=RESULT_BG, 
                             font=("Arial", 10), padx=5, pady=2)
            label.grid(row=row, column=0, sticky="w")
            
            value_label = tk.Label(result_details_frame, text=value, anchor="e", bg=RESULT_BG, 
                                  font=("Arial", 10, "bold"), padx=5, pady=2)
            value_label.grid(row=row, column=1, sticky="e")
            row += 1
            
        driver_label = tk.Label(result_details_frame, text="ИТОГО ВОДИТЕЛЮ:", anchor="w", 
                              bg=ACCENT_COLOR, fg="white", font=("Arial", 12, "bold"), padx=5, pady=3)
        driver_label.grid(row=row, column=0, sticky="we", pady=5)
        
        driver_value = tk.Label(result_details_frame, text=f"{driver_payment:.2f} руб.", anchor="e", 
                               bg=ACCENT_COLOR, fg="white", font=("Arial", 12, "bold"), padx=5, pady=3)
        driver_value.grid(row=row, column=1, sticky="we", pady=5)

        try:
            save_to_excel({
                "Дата": date_value,
                "ФИО": driver_fio,
                "Автомобиль": car_model,
                "Общая выручка": total_revenue,
                "Ручная комиссия": commission,
                "Трудовая выплата %": labor_percentage,
                "Трудовая выплата": labor_payment,
                "Расходы на бензин": gasoline,
                "Бонус за выходные": weekend_bonus,
                "Бонус за топ место %": top_place_bonus_percentage,
                "Бонус за топ место": top_place_bonus,
                "Выплата водителю": driver_payment,
                "Выплата парку": fleet_payment
            })
            status_label.config(text="✓ Данные успешно сохранены в Excel", fg="green")
        except Exception as e:
            status_label.config(text=f"❌ Ошибка сохранения: {str(e)}", fg="red")
            messagebox.showerror("Ошибка сохранения", f"Не удалось сохранить данные: {str(e)}")

    except ValueError:
        messagebox.showerror("Ошибка", "Введите корректные числа в поля 'Общая выручка' и 'Сумма комиссии'.")
    except Exception as e:
        messagebox.showerror("Ошибка", f"Произошла ошибка: {str(e)}")

def save_to_excel(data):
    try:
        if os.path.exists(EXCEL_FILENAME):
            try:
                with open(EXCEL_FILENAME, 'a+b') as f:
                    pass
            except PermissionError:
                raise PermissionError(f"Файл {EXCEL_FILENAME} уже открыт. Закройте его и повторите попытку.")
        
        new_file = False
        try:
            workbook = openpyxl.load_workbook(EXCEL_FILENAME)
        except FileNotFoundError:
            workbook = openpyxl.Workbook()
            if "Sheet" in workbook.sheetnames:
                sheet = workbook.active
                workbook.remove(sheet)
            new_file = True
        
        sheet_name = "Data"
        
        if sheet_name not in workbook.sheetnames:
            sheet = workbook.create_sheet(sheet_name)
            new_file = True
        else:
            sheet = workbook[sheet_name]
        
        headers_exist = False
        if sheet.max_row >= 1:
            first_row_values = [cell.value for cell in sheet[1]]
            required_headers = list(data.keys())
            
            if all(header in first_row_values for header in required_headers):
                headers_exist = True
        
        if not headers_exist or new_file:
            sheet.delete_rows(1, sheet.max_row)
            header = list(data.keys())
            sheet.append(header)
            
            for cell in sheet[1]:
                cell.font = openpyxl.styles.Font(bold=True)
                cell.fill = openpyxl.styles.PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
            
            print(">>> Заголовки добавлены в лист.")
        
        row_values = list(data.values())
        sheet.append(row_values)
        
        for column in sheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                if cell.value:
                    cell_length = len(str(cell.value))
                    if cell_length > max_length:
                        max_length = cell_length
            adjusted_width = (max_length + 2) * 1.2
            sheet.column_dimensions[column_letter].width = adjusted_width
        
        workbook.save(EXCEL_FILENAME)
        print("Данные сохранены в Excel.")
    except Exception as e:
        print(f"Ошибка при сохранении в Excel: {str(e)}")
        raise

def clear_excel_table():
    try:
        if os.path.exists(EXCEL_FILENAME):
            try:
                with open(EXCEL_FILENAME, 'a+b') as f:
                    pass
            except PermissionError:
                messagebox.showerror("Ошибка", f"Файл {EXCEL_FILENAME} уже открыт. Закройте его и повторите попытку.")
                return
                
        try:
            workbook = openpyxl.load_workbook(EXCEL_FILENAME)
            sheet_name = "Data"
            
            if sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                header = []
                if sheet.max_row > 0:
                    header = [cell.value for cell in sheet[1]]
                workbook.remove(sheet)
                
                new_sheet = workbook.create_sheet(sheet_name)
                
                if header:
                    new_sheet.append(header)
                    for cell in new_sheet[1]:
                        cell.font = openpyxl.styles.Font(bold=True)
                        cell.fill = openpyxl.styles.PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
            else:
                workbook.create_sheet(sheet_name)
                
            workbook.save(EXCEL_FILENAME)
            status_label.config(text="✓ Таблица Excel очищена", fg="green")
            messagebox.showinfo("Успех", "Таблица Excel успешно очищена.")
            print("Таблица Excel очищена.")
        except FileNotFoundError:
            status_label.config(text="❌ Файл Excel не найден", fg="red")
            messagebox.showinfo("Информация", "Файл Excel не найден. Будет создан при первом сохранении.")
            print("Файл Excel не найден.")
    except Exception as e:
        status_label.config(text=f"❌ Ошибка: {str(e)}", fg="red")
        messagebox.showerror("Ошибка", f"Произошла ошибка: {str(e)}")

def reset_form():
    fio_entry.delete(0, tk.END)
    date_var.set(datetime.date.today().strftime("%d.%m.%Y"))
    car_combobox.set(car_list[0])
    revenue_entry.delete(0, tk.END)
    commission_entry.delete(0, tk.END)
    weekend_var.set(False)
    top_place_var.set(0)
    result_label.config(text="")
    for widget in result_details_frame.winfo_children():
        widget.destroy()
    status_label.config(text="")
    result_frame.config(bg=MAIN_BG)

window = tk.Tk()
window.title("Калькулятор оплаты таксиста")
window.configure(bg=MAIN_BG)
window.geometry("1152x864")

style = ttk.Style()
style.configure("TButton", font=("Arial", 10), background=BUTTON_BG)
style.configure("TEntry", background=ENTRY_BG)
style.configure("TCombobox", background=ENTRY_BG)
style.configure("TCheckbutton", background=MAIN_BG)
style.configure("TRadiobutton", background=MAIN_BG)

header_frame = tk.Frame(window, bg=HEADER_BG, padx=10, pady=10)
header_frame.pack(fill=tk.X, padx=0, pady=0)

tk.Label(header_frame, text="КАЛЬКУЛЯТОР ОПЛАТЫ ТАКСИСТА", font=("Arial", 16, "bold"), 
         bg=HEADER_BG, fg=HEADER_FG).pack(pady=5)

main_frame = tk.Frame(window, bg=MAIN_BG, padx=10, pady=10)
main_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)

form_frame = tk.LabelFrame(main_frame, text="Данные водителя", bg=MAIN_BG, font=("Arial", 12), padx=10, pady=10)
form_frame.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=5, pady=5)

result_frame = tk.Frame(main_frame, bg=MAIN_BG, padx=10, pady=10)
result_frame.pack(side=tk.RIGHT, fill=tk.BOTH, expand=True, padx=5, pady=5)

vcmd_number = (window.register(validate_number_input), '%P')
vcmd_date = (window.register(validate_date_input), '%P')

tk.Label(form_frame, text="ФИО водителя:", bg=MAIN_BG, font=("Arial", 10)).grid(row=0, column=0, sticky='w', padx=5, pady=5)
fio_entry = tk.Entry(form_frame, width=30, font=("Arial", 10), bg=ENTRY_BG)
fio_entry.grid(row=0, column=1, padx=5, pady=5, sticky='w')

tk.Label(form_frame, text="Дата:", bg=MAIN_BG, font=("Arial", 10)).grid(row=1, column=0, sticky='w', padx=5, pady=5)

date_frame = tk.Frame(form_frame, bg=MAIN_BG)
date_frame.grid(row=1, column=1, padx=5, pady=5, sticky='w')

date_var = tk.StringVar()
date_var.set(datetime.date.today().strftime("%d.%m.%Y"))  # Set initial value to today

date_entry = tk.Entry(date_frame, textvariable=date_var, width=15, font=("Arial", 10), bg=ENTRY_BG, state="readonly")
date_entry.pack(side="left", padx=(0, 5))

date_button = tk.Button(date_frame, text="Выбрать", command=open_calendar, 
                      bg=BUTTON_BG, fg=BUTTON_FG, font=("Arial", 9))
date_button.pack(side="left")

car_list = [
    "P424EM27",
    "P853BA27",
    "A800EB49",
    "P865EO27",
    "A438BE49",
    "A504BT49",
    "A429EB49",
    "A123EE49",
    "A412EE49",
    "A197EB49",
    "A895EB49"
]
tk.Label(form_frame, text="Автомобиль:", bg=MAIN_BG, font=("Arial", 10)).grid(row=2, column=0, sticky='w', padx=5, pady=5)
car_combobox = ttk.Combobox(form_frame, values=car_list, width=28, font=("Arial", 10))
car_combobox.grid(row=2, column=1, padx=5, pady=5, sticky='w')
car_combobox.set(car_list[0])

ttk.Separator(form_frame, orient='horizontal').grid(row=3, column=0, columnspan=2, sticky='ew', pady=10)

tk.Label(form_frame, text="Финансовые данные", bg=MAIN_BG, font=("Arial", 12, "bold")).grid(row=4, column=0, columnspan=2, sticky='w', padx=5, pady=5)

tk.Label(form_frame, text="Общая выручка (руб.):", bg=MAIN_BG, font=("Arial", 10)).grid(row=5, column=0, sticky='w', padx=5, pady=5)
revenue_entry = tk.Entry(form_frame, width=30, font=("Arial", 10), bg=ENTRY_BG, validate="key", validatecommand=vcmd_number)
revenue_entry.grid(row=5, column=1, padx=5, pady=5, sticky='w')

tk.Label(form_frame, text="Сумма комиссии (руб.):", bg=MAIN_BG, font=("Arial", 10)).grid(row=6, column=0, sticky='w', padx=5, pady=5)
commission_entry = tk.Entry(form_frame, width=30, font=("Arial", 10), bg=ENTRY_BG, validate="key", validatecommand=vcmd_number)
commission_entry.grid(row=6, column=1, padx=5, pady=5, sticky='w')

ttk.Separator(form_frame, orient='horizontal').grid(row=7, column=0, columnspan=2, sticky='ew', pady=10)

tk.Label(form_frame, text="Дополнительные параметры", bg=MAIN_BG, font=("Arial", 12, "bold")).grid(row=8, column=0, columnspan=2, sticky='w', padx=5, pady=5)

weekend_var = tk.BooleanVar()
weekend_checkbox = tk.Checkbutton(form_frame, text="Водитель вышел в выходной день (+5%)", variable=weekend_var, 
                                 bg=MAIN_BG, font=("Arial", 10))
weekend_checkbox.grid(row=9, column=0, columnspan=2, sticky='w', padx=5, pady=5)

tk.Label(form_frame, text="Топ место водителя:", bg=MAIN_BG, font=("Arial", 10)).grid(row=10, column=0, sticky='w', padx=5, pady=5)

top_place_frame = tk.Frame(form_frame, bg=MAIN_BG)
top_place_frame.grid(row=11, column=0, columnspan=2, sticky='w', padx=5, pady=5)

top_place_var = tk.IntVar()
top_place_var.set(0)

top_place_radio_none = tk.Radiobutton(top_place_frame, text="Нет в топе", variable=top_place_var, value=0, 
                                     bg=MAIN_BG, font=("Arial", 10))
top_place_radio_1 = tk.Radiobutton(top_place_frame, text="1 место (+5%)", variable=top_place_var, value=1, 
                                  bg=MAIN_BG, font=("Arial", 10))
top_place_radio_2 = tk.Radiobutton(top_place_frame, text="2 место (+3%)", variable=top_place_var, value=2, 
                                  bg=MAIN_BG, font=("Arial", 10))
top_place_radio_3 = tk.Radiobutton(top_place_frame, text="3 место (+1%)", variable=top_place_var, value=3, 
                                  bg=MAIN_BG, font=("Arial", 10))

top_place_radio_none.grid(row=0, column=0, sticky='w')
top_place_radio_1.grid(row=0, column=1, sticky='w', padx=10)
top_place_radio_2.grid(row=1, column=0, sticky='w')
top_place_radio_3.grid(row=1, column=1, sticky='w', padx=10)

buttons_frame = tk.Frame(form_frame, bg=MAIN_BG)
buttons_frame.grid(row=12, column=0, columnspan=2, pady=15)

calculate_button = tk.Button(buttons_frame, text="Рассчитать и сохранить", command=calculate_payment_gui,
                            bg=BUTTON_BG, fg=BUTTON_FG, font=("Arial", 11, "bold"), padx=10, pady=5)
calculate_button.grid(row=0, column=0, padx=5)

reset_button = tk.Button(buttons_frame, text="Сбросить форму", command=reset_form,
                        bg="#e74c3c", fg=BUTTON_FG, font=("Arial", 11), padx=10, pady=5)
reset_button.grid(row=0, column=1, padx=5)

clear_table_button = tk.Button(buttons_frame, text="Очистить журнал Excel", command=clear_excel_table,
                             bg="#7f8c8d", fg=BUTTON_FG, font=("Arial", 11), padx=10, pady=5)
clear_table_button.grid(row=0, column=2, padx=5)

status_label = tk.Label(form_frame, text="", bg=MAIN_BG, font=("Arial", 9))
status_label.grid(row=13, column=0, columnspan=2, sticky='w', padx=5, pady=5)

result_label = tk.Label(result_frame, text="", bg=MAIN_BG, font=("Arial", 14, "bold"), anchor="w")
result_label.pack(fill=tk.X, pady=5)

result_details_frame = tk.Frame(result_frame, bg=MAIN_BG)
result_details_frame.pack(fill=tk.BOTH, expand=True)

footer_frame = tk.Frame(window, bg="#34495e", height=20)
footer_frame.pack(fill=tk.X, padx=0, pady=0)
tk.Label(footer_frame, text="© Калькулятор оплаты таксиста v1.0", bg="#34495e", fg="white", 
         font=("Arial", 8)).pack(side=tk.RIGHT, padx=10, pady=3)

window.mainloop()